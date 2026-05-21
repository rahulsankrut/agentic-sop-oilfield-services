"""Load EIA Short-Term Energy Outlook data into BigQuery.

TASK-16 Step 4c.

Pulls the STEO bulk Excel (no API key needed — public, US Govt domain)
from https://www.eia.gov/outlooks/steo/xls/STEO_m.xlsx and writes the
monthly basin-level rig count + tight-oil + shale-gas series into
`eia_steo.basin_production`.

STEO uses two overlapping classification schemes:
  - Table 10a (Drilling Productivity Metrics) — "regions": Appalachia,
    Bakken, Eagle Ford, Haynesville, Permian, Rest of L48.
  - Table 10b (Tight oil + shale gas production) — "formations": Austin
    Chalk, Bakken, Eagle Ford, Haynesville, Marcellus, Mississippian,
    Niobrara Codell, Permian, Utica, Woodford, etc.

Same name overlap (Bakken/Eagle Ford/Permian/Haynesville/Appalachia)
gets merged into one row per (BASIN, REPORT_MONTH). Non-overlapping
names get their own rows — ACTIVE_RIGS NULL for formations not in 10a,
production NULL for regions not in 10b.

Substitution: this dataset is the same for everyone (US Govt public
domain). Refresh by re-running this script. EIA releases STEO ~monthly,
typically the 2nd Thursday of each month.

This intentionally skips the EIA Open Data API v2 — that requires an
API key (free signup, but adds a setup step). The STEO bulk Excel
covers the same data we need without auth, and is what production
customers would use anyway (small file, easy to wire into Cloud
Scheduler).

Attribution: "Source: U.S. Energy Information Administration,
Short-Term Energy Outlook"
"""

from __future__ import annotations

import argparse
import calendar
import logging
import sys
import urllib.request
from datetime import UTC, date, datetime

UTC = UTC  # Python 3.10 compat (datetime.UTC is 3.11+)
from pathlib import Path

from google.cloud import bigquery
from openpyxl import load_workbook

PROJECT = "vertex-ai-demos-468803"
TABLE = "eia_steo.basin_production"
AUDIT_TABLE = "bakerhughes_rig_count.dataset_loads"
STEO_URL = "https://www.eia.gov/outlooks/steo/xls/STEO_m.xlsx"

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Code → BASIN name mappings (STEO uses opaque codes; we map to friendly names
# that match what an OFS planner would say. BASIN column is STRING(20).)
# ---------------------------------------------------------------------------

# Table 10a — Active rigs by region.
RIG_REGIONS = {
    "RIGSAP": "Appalachia",
    "RIGSBK": "Bakken",
    "RIGSEF": "Eagle Ford",
    "RIGSHA": "Haynesville",
    "RIGSPM": "Permian",
    "RIGSR48": "Rest of L48",
}

# Table 10b — Tight oil production by formation (million bbl/day → bbl/day).
OIL_FORMATIONS = {
    "TOPRAC": "Austin Chalk",
    "TOPRBK": "Bakken",
    "TOPREF": "Eagle Ford",
    "TOPRMP": "Mississippian",
    "TOPRNI": "Niobrara",
    "TOPRPM": "Permian",
    "TOPRWF": "Woodford",
}

# Table 10b — Shale dry gas production by formation (Bcf/day → Mcf/day).
GAS_FORMATIONS = {
    "SNGPRBK": "Bakken",
    "SNGPRBN": "Barnett",
    "SNGPREF": "Eagle Ford",
    "SNGPRFY": "Fayetteville",
    "SNGPRHA": "Haynesville",
    "SNGPRMC": "Marcellus",
    "SNGPRMP": "Mississippian",
    "SNGPRNI": "Niobrara",
    "SNGPRPM": "Permian",
    "SNGPRUA": "Utica",
    "SNGPRWF": "Woodford",
}


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------


def _fetch_steo(url: str, out: Path) -> None:
    log.info("downloading STEO bulk Excel from %s ...", url)
    req = urllib.request.Request(url, headers={"User-Agent": "oilfield-services-task16/1.0"})
    with urllib.request.urlopen(req, timeout=60) as resp, out.open("wb") as f:
        f.write(resp.read())
    log.info("downloaded %.2f MiB", out.stat().st_size / 1024 / 1024)


def _month_columns(sheet) -> list[date]:
    """Build the ordered list of (year, month) dates for every data column.

    Layout: row 3 has year integers spaced out (one cell per year), row 4
    has month names (Jan..Dec) in the data columns. We walk the year row
    forward-filling years until the next year cell appears.
    """
    year_row = list(next(sheet.iter_rows(min_row=3, max_row=3, values_only=True)))
    month_row = list(next(sheet.iter_rows(min_row=4, max_row=4, values_only=True)))
    months_by_name = {m: i for i, m in enumerate(calendar.month_abbr) if m}
    out: list[date | None] = []
    current_year: int | None = None
    for i, m in enumerate(month_row):
        if isinstance(year_row[i], int):
            current_year = year_row[i]
        if isinstance(m, str) and m in months_by_name and current_year is not None:
            out.append(date(current_year, months_by_name[m], 1))
        else:
            out.append(None)
    return out


def _parse_series(sheet, code_to_basin: dict[str, str]) -> dict[tuple[str, date], float]:
    """Read a sheet and return {(basin, month_date): value} for the codes we want."""
    months = _month_columns(sheet)
    out: dict[tuple[str, date], float] = {}
    for row in sheet.iter_rows(min_row=5, values_only=True):
        code = row[0]
        if not isinstance(code, str) or code not in code_to_basin:
            continue
        basin = code_to_basin[code]
        for col_idx, cell in enumerate(row):
            month = months[col_idx] if col_idx < len(months) else None
            if month is None or not isinstance(cell, (int, float)):
                continue
            out[(basin, month)] = float(cell)
    return out


def _build_rows(steo_path: Path) -> list[dict]:
    log.info("parsing STEO Excel %s ...", steo_path)
    wb = load_workbook(steo_path, read_only=True, data_only=True)
    rigs = _parse_series(wb["10atab"], RIG_REGIONS)
    oil = _parse_series(wb["10btab"], OIL_FORMATIONS)
    gas = _parse_series(wb["10btab"], GAS_FORMATIONS)
    log.info("parsed 10a rigs: %d series points", len(rigs))
    log.info("parsed 10b oil:  %d series points", len(oil))
    log.info("parsed 10b gas:  %d series points", len(gas))

    keys = set(rigs.keys()) | set(oil.keys()) | set(gas.keys())
    rows = []
    for basin, month in sorted(keys, key=lambda k: (k[1], k[0])):
        oil_bpd = oil.get((basin, month))
        gas_bcfd = gas.get((basin, month))
        rows.append(
            {
                "REPORT_MONTH": month.isoformat(),
                "BASIN": basin[:20],
                # 10a rigs are monthly averages — fractional is intentional, preserved.
                "ACTIVE_RIGS": rigs.get((basin, month)),
                # 10b tight oil ships in million bbl/day → convert to bbl/day (int).
                "OIL_PROD_BPD": int(oil_bpd * 1_000_000) if oil_bpd is not None else None,
                # 10b shale gas ships in Bcf/day → convert to Mcf/day (×1,000,000).
                "GAS_PROD_MCFD": int(gas_bcfd * 1_000_000) if gas_bcfd is not None else None,
                # STEO doesn't include the DPR's new-well productivity series anymore.
                "RIG_PRODUCTIVITY_NEW_WELL_OIL_BPD": None,
            }
        )
    log.info("built %d (basin, month) rows", len(rows))
    return rows


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------


def _audit_row(source: str, row_count: int, started: datetime, ended: datetime) -> dict:
    return {
        "dataset": "eia_steo.basin_production",
        "source_url": source,
        "row_count": row_count,
        "load_started_at": started.isoformat(),
        "load_completed_at": ended.isoformat(),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--file",
        type=Path,
        help="Path to a local STEO_m.xlsx (skips the EIA download).",
    )
    p.add_argument(
        "--url",
        default=STEO_URL,
        help="Override the EIA STEO bulk Excel URL.",
    )
    args = p.parse_args()

    started = datetime.now(tz=UTC)

    if args.file:
        steo_path = args.file
        source = str(args.file)
    else:
        steo_path = Path("/tmp/eia_steo_m.xlsx")
        _fetch_steo(args.url, steo_path)
        source = args.url

    rows = _build_rows(steo_path)
    if not rows:
        log.error("no rows parsed from STEO")
        return 2

    client = bigquery.Client(project=PROJECT)
    job = client.load_table_from_json(
        rows,
        f"{PROJECT}.{TABLE}",
        job_config=bigquery.LoadJobConfig(
            write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE,
            source_format=bigquery.SourceFormat.NEWLINE_DELIMITED_JSON,
        ),
    )
    job.result()
    ended = datetime.now(tz=UTC)
    log.info("loaded %d rows into %s", job.output_rows or 0, TABLE)

    audit_job = client.load_table_from_json(
        [_audit_row(source, job.output_rows or 0, started, ended)],
        f"{PROJECT}.{AUDIT_TABLE}",
        job_config=bigquery.LoadJobConfig(
            write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
            source_format=bigquery.SourceFormat.NEWLINE_DELIMITED_JSON,
        ),
    )
    audit_job.result()
    log.info("audit row appended to %s", AUDIT_TABLE)
    return 0


if __name__ == "__main__":
    sys.exit(main())
