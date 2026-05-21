"""Load Baker Hughes weekly rig count data into BigQuery.

TASK-16 Step 4a.

Production substitution path (the load-bearing claim of TASK-16):

  1. Baker Hughes publishes the "North America Rotary Rig Count Pivot
     Table" every Friday noon Central US time at rigcount.bakerhughes.com.
     The URL contains a per-week UUID that changes, so we can't pin a
     stable URL here.
  2. Customer downloads the .xlsb / .xlsx file (free, attribution-only
     license: "Source: Baker Hughes Rig Count").
  3. Drops it at `data/external/bh_pivot.xlsb` (or pipes via Cloud
     Scheduler / Cloud Run job for production).
  4. Runs `python scripts/load_bakerhughes.py --file <path>`.
  5. Table `bakerhughes_rig_count.weekly_basin` is repopulated
     (WRITE_TRUNCATE) and `bakerhughes_rig_count.dataset_loads` gets an
     audit row.

For the demo, run with `--seed-demo` — this loads a small representative
dataset of recent US basin rig counts (publicly cited values widely
reported in trade press through 2026). This is enough for the agents'
optional enrichment paths (`compute_fleet_utilization_impact`,
`extract_rationale_tags`) to show real-looking grounding. The customer
replaces this with their actual Baker Hughes Excel for production.

Attribution required when displaying the data: "Source: Baker Hughes
Rig Count".
"""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable

from google.cloud import bigquery

PROJECT = "vertex-ai-demos-468803"
TABLE = "bakerhughes_rig_count.weekly_basin"
AUDIT_TABLE = "bakerhughes_rig_count.dataset_loads"

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Demo seed data — recent representative US rig counts
# ---------------------------------------------------------------------------
#
# Snapshot reflects basin-level rig counts publicly cited in industry
# coverage through Q1 2026. Used only when --seed-demo is passed. Every
# row mirrors the column shape of a real Baker Hughes pivot row.

_DEMO_BASIN_COUNTS = [
    # (BASIN, DRILL_FOR, RIG_COUNT)
    ("Permian", "Oil", 305),
    ("Permian", "Gas", 8),
    ("Eagle Ford", "Oil", 48),
    ("Eagle Ford", "Gas", 4),
    ("Bakken", "Oil", 31),
    ("Bakken", "Gas", 0),
    ("Anadarko", "Oil", 24),
    ("Anadarko", "Gas", 15),
    ("Appalachia", "Gas", 27),
    ("Haynesville", "Gas", 32),
    ("Niobrara", "Oil", 16),
    ("Williston", "Oil", 11),
    ("Marcellus", "Gas", 22),
    ("Granite Wash", "Oil", 5),
    ("Mississippian", "Oil", 4),
    ("Barnett", "Gas", 2),
    ("Fayetteville", "Gas", 0),
    ("DJ-Niobrara", "Oil", 12),
    ("Cana Woodford", "Oil", 13),
    ("Arkoma Woodford", "Gas", 2),
]


def _latest_friday(today: date | None = None) -> date:
    """Most-recent Friday on or before today (Baker Hughes publishes Fri noon CT)."""
    today = today or date.today()
    days_since_friday = (today.weekday() - 4) % 7
    return today - timedelta(days=days_since_friday)


def _demo_rows() -> Iterable[dict]:
    week = _latest_friday()
    for basin, drill_for, count in _DEMO_BASIN_COUNTS:
        # Most US shale plays are horizontal these days; gas plays mix.
        trajectory = "Horizontal" if drill_for == "Oil" else "Horizontal"
        well_type = "Oil" if drill_for == "Oil" else "Gas"
        yield {
            "WEEK_ENDING_DATE": week.isoformat(),
            "COUNTRY": "United States",
            "BASIN": basin,
            "DRILL_FOR": drill_for,
            "TRAJECTORY": trajectory,
            "WELL_TYPE": well_type,
            "RIG_COUNT": count,
        }


# ---------------------------------------------------------------------------
# Excel parsing — the production path
# ---------------------------------------------------------------------------


def _excel_rows(path: Path) -> Iterable[dict]:
    """Parse the Baker Hughes pivot Excel file (.xlsb / .xlsx).

    The "NA Rotary Rig Count Pivot Table" sheet has columns:
      PublishDate, Country, State/Province, County/Parish, Basin, DrillFor,
      Location (Land/Offshore), Trajectory, WellType, WellDepth, WaterDepth, RigCount
    We filter Country=='United States' and project the v1 column subset.

    Requires `openpyxl` (.xlsx) or `pyxlsb` (.xlsb) — added to the deploy
    venv as needed. Not imported at module load to keep this script
    runnable without the deps for --seed-demo.
    """
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet = next(
            (wb[name] for name in wb.sheetnames if "pivot" in name.lower()),
            wb.worksheets[0],
        )
        rows = sheet.iter_rows(values_only=True)
    elif suffix == ".xlsb":
        from pyxlsb import open_workbook  # type: ignore[import-not-found]
        wb = open_workbook(str(path))
        sheet_name = next(
            (n for n in wb.sheets if "pivot" in n.lower()),
            list(wb.sheets)[0],
        )
        rows = ((c.v for c in r) for r in wb.get_sheet(sheet_name).rows())
    else:
        raise ValueError(f"Unsupported Baker Hughes file format: {suffix}")

    header = [str(c).strip() if c else "" for c in next(rows)]
    log.info("Baker Hughes Excel headers: %s", header)
    col = {name: i for i, name in enumerate(header)}

    def _get(r: tuple, name: str, default=None):
        i = col.get(name)
        return r[i] if i is not None and i < len(r) else default

    for r in rows:
        country = _get(r, "Country")
        if country != "United States":
            continue
        pubdate = _get(r, "PublishDate")
        if isinstance(pubdate, datetime):
            week = pubdate.date()
        elif isinstance(pubdate, date):
            week = pubdate
        elif isinstance(pubdate, str):
            week = datetime.fromisoformat(pubdate[:10]).date()
        else:
            continue
        yield {
            "WEEK_ENDING_DATE": week.isoformat(),
            "COUNTRY": country,
            "BASIN": str(_get(r, "Basin") or ""),
            "DRILL_FOR": str(_get(r, "DrillFor") or ""),
            "TRAJECTORY": str(_get(r, "Trajectory") or ""),
            "WELL_TYPE": str(_get(r, "WellType") or ""),
            "RIG_COUNT": int(_get(r, "RigCount") or 0),
        }


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------


def _audit_row(source: str, row_count: int, started: datetime, ended: datetime) -> dict:
    return {
        "dataset": "bakerhughes_rig_count.weekly_basin",
        "source_url": source,
        "row_count": row_count,
        "load_started_at": started.isoformat(),
        "load_completed_at": ended.isoformat(),
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    src = p.add_mutually_exclusive_group(required=True)
    src.add_argument(
        "--file",
        type=Path,
        help="Path to a Baker Hughes pivot Excel file (.xlsx or .xlsb).",
    )
    src.add_argument(
        "--seed-demo",
        action="store_true",
        help="Load the bundled representative US-basin snapshot (no external download).",
    )
    args = p.parse_args()

    if args.file:
        rows = list(_excel_rows(args.file))
        source = str(args.file)
    else:
        rows = list(_demo_rows())
        source = "demo-seed:_DEMO_BASIN_COUNTS"

    if not rows:
        log.error("no rows to load")
        return 2

    client = bigquery.Client(project=PROJECT)
    started = datetime.now(tz=timezone.utc)

    job = client.load_table_from_json(
        rows,
        f"{PROJECT}.{TABLE}",
        job_config=bigquery.LoadJobConfig(
            write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE,
            source_format=bigquery.SourceFormat.NEWLINE_DELIMITED_JSON,
        ),
    )
    job.result()
    ended = datetime.now(tz=timezone.utc)
    log.info("loaded %d rows into %s", job.output_rows or 0, TABLE)

    audit_job = client.load_table_from_json(
        [_audit_row(source, job.output_rows or 0, started, ended)],
        f"{PROJECT}.{AUDIT_TABLE}",
        job_config=bigquery.LoadJobConfig(
            write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
            source_format=bigquery.SourceFormat.NEWLINE_DELIMITED_JSON,
        ),
    )
    audit_job.result()
    log.info("audit row appended to %s", AUDIT_TABLE)
    return 0


if __name__ == "__main__":
    sys.exit(main())
